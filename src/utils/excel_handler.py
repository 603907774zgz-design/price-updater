"""
Excel处理工具
"""
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from src.models.database import db, StandardSKU, PendingUpdate, UpdateSession


class ExcelHandler:
    """Excel处理器"""
    
    # 颜色定义
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    PRICE_UP_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 红色-涨价
    PRICE_DOWN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 绿色-降价
    PRICE_SAME_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # 白色-不变
    NEED_CONFIRM_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # 黄色-需确认
    
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    NORMAL_FONT = Font(size=10)
    
    CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
    LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
    
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self):
        self.session = db.get_session()
    
    def generate_update_excel(self, session_id: str, output_path: str = None) -> str:
        """
        生成更新结果Excel
        
        Args:
            session_id: 更新会话ID
            output_path: 输出路径（可选）
            
        Returns:
            生成的文件路径
        """
        # 获取待更新项
        pending_items = self.session.query(PendingUpdate).filter_by(
            session_id=session_id
        ).all()
        
        if not output_path:
            output_path = f"/tmp/price_update_{session_id}.xlsx"
        
        # 创建工作簿
        wb = Workbook()
        
        # Sheet 1: 变动明细
        ws1 = wb.active
        ws1.title = "价格变动明细"
        self._fill_update_sheet(ws1, pending_items)
        
        # Sheet 2: 需确认项
        ws2 = wb.create_sheet("需确认项")
        self._fill_confirm_sheet(ws2, pending_items)
        
        # Sheet 3: 新SKU
        ws3 = wb.create_sheet("新SKU提醒")
        self._fill_new_sku_sheet(ws3, pending_items)
        
        # Sheet 4: 完整标准表
        ws4 = wb.create_sheet("完整标准表")
        self._fill_full_standard_sheet(ws4)
        
        # 保存
        wb.save(output_path)
        
        return output_path
    
    def _fill_update_sheet(self, ws, pending_items):
        """填充变动明细表"""
        # 标题行
        headers = ["SKU编码", "商品分类", "商品系列", "商品标题", "规格", "颜色", 
                   "原价格", "新价格", "变动", "状态", "原始文本"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.CENTER_ALIGN
            cell.border = self.THIN_BORDER
        
        # 数据行
        row = 2
        for item in pending_items:
            if not item.matched_sku_code:
                continue
            
            sku = self.session.query(StandardSKU).filter_by(
                sku_code=item.matched_sku_code
            ).first()
            
            if not sku:
                continue
            
            # 确定状态和价格变动
            if item.match_status == 'CONFIRMED':
                status = "已确认"
            elif item.match_status == 'REJECTED':
                status = "已拒绝"
            else:
                status = "待确认"
            
            old_price = sku.price or 0
            new_price = item.extracted_price or 0
            change = new_price - old_price
            
            # 填充数据
            ws.cell(row=row, column=1, value=sku.sku_code)
            ws.cell(row=row, column=2, value=sku.category)
            ws.cell(row=row, column=3, value=sku.series)
            ws.cell(row=row, column=4, value=sku.title)
            ws.cell(row=row, column=5, value=sku.spec)
            ws.cell(row=row, column=6, value=sku.color)
            ws.cell(row=row, column=7, value=old_price)
            ws.cell(row=row, column=8, value=new_price)
            ws.cell(row=row, column=9, value=change if change != 0 else "-")
            ws.cell(row=row, column=10, value=status)
            ws.cell(row=row, column=11, value=item.raw_text)
            
            # 设置样式
            for col in range(1, 12):
                cell = ws.cell(row=row, column=col)
                cell.border = self.THIN_BORDER
                cell.font = self.NORMAL_FONT
                
                if col in [7, 8, 9]:  # 价格列居中
                    cell.alignment = self.CENTER_ALIGN
                else:
                    cell.alignment = self.LEFT_ALIGN
                
                # 根据价格变动设置背景色
                if col == 9:  # 变动列
                    if change > 0:
                        cell.fill = self.PRICE_UP_FILL
                    elif change < 0:
                        cell.fill = self.PRICE_DOWN_FILL
                
                # 待确认项高亮
                if item.match_status == 'PENDING':
                    cell.fill = self.NEED_CONFIRM_FILL
            
            row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 10
        ws.column_dimensions['J'].width = 10
        ws.column_dimensions['K'].width = 30
    
    def _fill_confirm_sheet(self, ws, pending_items):
        """填充需确认表"""
        headers = ["ID", "原始文本", "提取品牌", "提取系列", "提取规格", "提取颜色", 
                   "提取价格", "预激活", "匹配SKU", "匹配分数", "建议操作"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.CENTER_ALIGN
            cell.border = self.THIN_BORDER
        
        row = 2
        for item in pending_items:
            if item.match_status != 'PENDING':
                continue
            
            ws.cell(row=row, column=1, value=item.id)
            ws.cell(row=row, column=2, value=item.raw_text)
            ws.cell(row=row, column=3, value=item.extracted_brand)
            ws.cell(row=row, column=4, value=item.extracted_series)
            ws.cell(row=row, column=5, value=item.extracted_spec)
            ws.cell(row=row, column=6, value=item.extracted_color)
            ws.cell(row=row, column=7, value=item.extracted_price)
            ws.cell(row=row, column=8, value="是" if item.is_preactivated else "否")
            ws.cell(row=row, column=9, value=item.matched_sku_code)
            ws.cell(row=row, column=10, value=item.match_score)
            
            # 建议操作
            if item.match_score >= 75:
                suggestion = "建议确认"
            else:
                suggestion = "建议人工核对"
            ws.cell(row=row, column=11, value=suggestion)
            
            # 样式
            for col in range(1, 12):
                cell = ws.cell(row=row, column=col)
                cell.border = self.THIN_BORDER
                cell.font = self.NORMAL_FONT
                cell.fill = self.NEED_CONFIRM_FILL
            
            row += 1
        
        # 调整列宽
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            ws.column_dimensions[col].width = 15
        ws.column_dimensions['B'].width = 30
    
    def _fill_new_sku_sheet(self, ws, pending_items):
        """填充新SKU提醒表"""
        headers = ["原始文本", "提取品牌", "提取系列", "提取规格", "提取颜色", "提取价格", "备注"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.CENTER_ALIGN
            cell.border = self.THIN_BORDER
        
        row = 2
        for item in pending_items:
            if item.matched_sku_code:
                continue  # 只显示未匹配的
            
            ws.cell(row=row, column=1, value=item.raw_text)
            ws.cell(row=row, column=2, value=item.extracted_brand)
            ws.cell(row=row, column=3, value=item.extracted_series)
            ws.cell(row=row, column=4, value=item.extracted_spec)
            ws.cell(row=row, column=5, value=item.extracted_color)
            ws.cell(row=row, column=6, value=item.extracted_price)
            ws.cell(row=row, column=7, value="标准库中未找到匹配SKU，请手动添加")
            
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.border = self.THIN_BORDER
                cell.font = self.NORMAL_FONT
            
            row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 30
        for col in ['B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 15
    
    def _fill_full_standard_sheet(self, ws):
        """填充完整标准表"""
        # 获取所有SKU
        skus = self.session.query(StandardSKU).all()
        
        headers = ["SKU编码", "商品分类", "商品系列", "商品标题", "商品规格", "商品颜色", "商品行情价"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.CENTER_ALIGN
            cell.border = self.THIN_BORDER
        
        row = 2
        for sku in skus:
            ws.cell(row=row, column=1, value=sku.sku_code)
            ws.cell(row=row, column=2, value=sku.category)
            ws.cell(row=row, column=3, value=sku.series)
            ws.cell(row=row, column=4, value=sku.title)
            ws.cell(row=row, column=5, value=sku.spec)
            ws.cell(row=row, column=6, value=sku.color)
            ws.cell(row=row, column=7, value=sku.price)
            
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.border = self.THIN_BORDER
                cell.font = self.NORMAL_FONT
                
                if sku.price is None:
                    cell.fill = self.NEED_CONFIRM_FILL  # 空价格标黄
            
            row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
    
    def generate_standard_excel(self, output_path: str = None) -> str:
        """
        生成标准价格表Excel（当前最新价格）
        
        Returns:
            生成的文件路径
        """
        # 获取所有SKU
        skus = self.session.query(StandardSKU).all()
        
        if not output_path:
            output_path = f"/tmp/standard_prices_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "标准价格表"
        
        # 标题行
        headers = ["商品分类", "商品系列", "商品标题", "商品规格", "商品颜色", "商品行情价", "sku编码"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.CENTER_ALIGN
            cell.border = self.THIN_BORDER
        
        # 数据行
        for row, sku in enumerate(skus, 2):
            ws.cell(row=row, column=1, value=sku.category)
            ws.cell(row=row, column=2, value=sku.series)
            ws.cell(row=row, column=3, value=sku.title)
            ws.cell(row=row, column=4, value=sku.spec)
            ws.cell(row=row, column=5, value=sku.color)
            ws.cell(row=row, column=6, value=sku.price)
            ws.cell(row=row, column=7, value=sku.sku_code)
            
            # 设置样式
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.border = self.THIN_BORDER
                cell.font = self.NORMAL_FONT
                
                if col == 6 and sku.price is None:
                    # 空价格标黄
                    cell.fill = self.NEED_CONFIRM_FILL
        
        # 调整列宽
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        
        wb.save(output_path)
        
        return output_path

    def close(self):
        """关闭会话"""
        self.session.close()


if __name__ == '__main__':
    handler = ExcelHandler()
    # 测试生成
    # handler.generate_update_excel("test_session_id")
    handler.close()
